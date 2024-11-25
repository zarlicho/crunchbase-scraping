from selenium.webdriver.common.by import By
from seleniumbase import Driver
from dotenv import load_dotenv
from lxml import html
import openpyxl,os

load_dotenv()
class Cnbase:
	def __init__(self):
		self.sb = Driver(uc=True,headless=False)
		self.email = os.getenv('EMAIL')
		self.password = os.getenv('PASS')

	def login(self):
		self.sb.uc_open_with_reconnect("https://www.crunchbase.com/login", 6.111)
		global status
		status = False
		for retry in range(5):
			try:
				if self.sb.is_text_visible(text="Oops! There was a problem!",selector='//*[@id="mat-mdc-dialog-title-0"]',by=By.XPATH):
					self.sb.refresh()
					continue
				self.sb.send_keys('/html/body/chrome/div/mat-sidenav-container/mat-sidenav-content/div/authentication-page/page-layout/div/div/div[2]/authentication/mat-card/mat-tab-group/div/mat-tab-body[1]/div/login/form/mat-form-field[1]/div[1]/div/div[2]/input', os.getenv('EMAIL'),by=By.XPATH,timeout=20)
				self.sb.send_keys('/html/body/chrome/div/mat-sidenav-container/mat-sidenav-content/div/authentication-page/page-layout/div/div/div[2]/authentication/mat-card/mat-tab-group/div/mat-tab-body[1]/div/login/form/mat-form-field[2]/div[1]/div/div[2]/input', os.getenv('PASS'),by=By.XPATH,timeout=20)
				self.sb.uc_click(selector='/html/body/chrome/div/mat-sidenav-container/mat-sidenav-content/div/authentication-page/page-layout/div/div/div[2]/authentication/mat-card/mat-tab-group/div/mat-tab-body[1]/div/login/form/div[2]/button[1]', by=By.XPATH,timeout=20)
				if self.sb.is_valid_url("https://www.crunchbase.com/home"):
					status = True
					break
				else:
					status = False
					continue
			except Exception as e:
				print(f"trying in {retry+1}x")
				self.sb.refresh()
		return status

		# time.sleep(100)
	def storeData(self,filename, data):
		if os.path.exists(filename):
			workbook = openpyxl.load_workbook(filename)
			sheet = workbook.active
			start_row = sheet.max_row + 1
		else:
			workbook = openpyxl.Workbook()
			sheet = workbook.active
			sheet.append(['Investor', 'Companies', 'Website', 'Linkedin'])
			start_row = 2
		for item in data:
			investor = item['Investor']
			companies = item['Companies']
			website = item['Website']
			linkedin = item['Linkedin']
			for x,y in enumerate(companies):
				sheet.append([investor, y, website[x], linkedin[x]])
			sheet.append(['', '', '', ''])
		workbook.save(filename)
		print(f"Data has been appended to '{filename}'.")

	def get_element_bs4(self,selector,pgSource,types=""):
		tree = html.fromstring(pgSource)
		if len(tree.xpath(selector)) != 0:
			if types=="text":
				return tree.xpath(selector)[0].text_content()
			else:
				return tree.xpath(selector)[0]
		else:
			return None

	def get_elements(self, selector, by):
		try:
			self.sb.wait_for_element_visible(selector, by=by, timeout=10)
			if self.sb.is_element_visible(selector, by):
				try:
					return self.sb.find_elements(selector, by, limit=50)
				except Exception as e:
					print(f"Error finding elements {selector}: {e}")
					return None
			else:
				print(f"Element {selector} not found!")
				return None
		except Exception as e:
			print(f"Exception in get_elements for {selector}: {e}")
			return None

	def is_element_nill(self,selector): # find elements with check element first
		if self.sb.is_element_present(selector=selector,by=By.CSS_SELECTOR):
			return self.sb.find_element(selector=selector,by=By.CSS_SELECTOR,timeout=5)
		else:
			return None

	def getProfile(self,profUrl):
		self.sb.open(profUrl)
		investName = self.get_element_bs4(selector="/html/body/chrome/div/mat-sidenav-container/mat-sidenav-content/div/ng-component/entity-v2/page-layout/div/div/profile-header/div/header/div/div/div/div[2]/div[1]/h1/text()",pgSource=self.sb.get_page_source())
		investWeb = self.is_element_nill('a[_ngcontent-ng-c699674806]')
		investLinkedin = self.sb.find_elements(selector="//*[@class='icon']", by="xpath")
		if not investLinkedin:
			return investName, investWeb.get_attribute("href") if investWeb else "not found", "not found"
		all_links = [
			a.get_attribute("href")
			for icon_element in investLinkedin
			for li in icon_element.find_elements(By.TAG_NAME, "li")
			for a in li.find_elements(By.TAG_NAME, "a")
			if "linkedin" in a.get_attribute("href")
		]
		linkedin_link = all_links[0] if all_links else "not found"
		website_link = investWeb.get_attribute("href") if investWeb else "not found"

		return investName, website_link, linkedin_link

 
	def getInvestments(self,investmentUrl):
		self.sb.open(investmentUrl+"/recent_investments/investments")
		self.sb.uc_click(selector='//*[@id="mat-tab-nav-panel-0"]/div/div/page-centered-layout[2]/div/profile-section/section-card/mat-card/div[2]/list-card/list-card-more-results/button',by=By.XPATH)
		orgaName = self.sb.find_elements(selector='//*[@id="mat-tab-nav-panel-0"]/div/div/page-centered-layout[2]/div/profile-section/section-card/mat-card/div[2]/list-card/div/table/tbody',by=By.XPATH)
		investmentName = self.sb.find_elements(selector='/html/body/chrome/div/mat-sidenav-container/mat-sidenav-content/div/ng-component/entity-v2/page-layout/div/div/profile-header/div/header/div/div/div/div[2]/div[1]/h1',by=By.XPATH)[0].text.split("\n")[0]
		companyNames = []
		companyWebs = []
		companyLinkedins = []
		for tbody in orgaName:
			all_organization = [(tra.get_attribute("href")) for tr in tbody.find_elements(By.TAG_NAME,"tr") for tra in tr.find_elements(By.TAG_NAME,"a") if "organization" in tra.get_attribute("href")]
			for orgaurl in all_organization:
				Name,Web,linkedin = self.getProfile(orgaurl)
				print(Name,Web,linkedin)
				if Name and Web and linkedin:
					companyWebs.append(Web)
					companyLinkedins.append(linkedin)
					companyNames.append(Name)
		new_data = [
			{
				'Investor':investmentName,
				'Companies':[name for name in companyNames],
				'Website':[webs for webs in companyWebs],
				'Linkedin':[linked for linked in companyLinkedins]
			}
		]
		self.storeData('./data.xlsx',new_data)

if __name__ == "__main__":
	crunchbase = Cnbase()
	if crunchbase.login():
		print("login successfully!")
		# crunchbase.getInvestments("https://www.crunchbase.com/organization/nashville-entrepreneur-center")
